from flask import Flask, make_response, render_template, request, redirect, url_for, session, flash, jsonify
from config import db  # N'importez plus 'auth' depuis config
from firebase_admin import auth  # Utilisez uniquement Firebase Admin
from firebase_admin.exceptions import FirebaseError
from datetime import datetime
import requests

app = Flask(__name__)
app.secret_key = "supersecretkeySIS2025"

def is_valid_date(date_str):
    try:
        datetime.fromisoformat(date_str)
        return True
    except ValueError:
        return False

# Page de connexion
# @app.route("/", methods=["GET", "POST"])
# def login():
#     if request.method == "POST":
#         email = request.form["username"]
#         password = request.form["password"]
#         try:
#             user = auth.sign_in_with_email_and_password(email, password)
#             session["user"] = user['localId']
#             return redirect(url_for("home"))
#         except Exception as e:
#             flash(f"Identifiants incorrects: {str(e)}")
#             return redirect(url_for("login"))
#     return render_template("login.html")
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        
        if not email or not password:
            flash("Email et mot de passe requis", 'error')
            return redirect(url_for("login"))

        try:
            # Solution 1: Authentification avec Admin SDK (sans vérification mot de passe)
            user = auth.get_user_by_email(email)
            
            # Vérification email (optionnelle)
            if not user.email_verified:
                flash("Veuillez vérifier votre email d'abord", 'error')
                return redirect(url_for("login"))
                
            session["user"] = {
                'uid': user.uid,
                'email': user.email,
                'display_name': user.email.split('@')[0]
            }
            return redirect(url_for("home"))
            
        except auth.UserNotFoundError:
            flash("Email incorrect", 'error')
        except Exception as e:
            print(f"Erreur connexion: {str(e)}")
            flash("Erreur technique lors de la connexion", 'error')
        
        return redirect(url_for("login"))
    
    return render_template("login.html")
# Page d'accueil
@app.route("/home")
def home():
    if "user" not in session:
        return redirect(url_for("login"))
    
    members_ref = db.collection("members")
    members = members_ref.stream()
    members_list = [{"id": doc.id, **doc.to_dict()} for doc in members]
    
    return render_template("home.html", 
                        members=members_list,
                        year=request.args.get("year", ""),
                        month=request.args.get("month", ""),
                        city=request.args.get("city", ""))

# Ajouter un membre
@app.route("/add_member", methods=["POST"])
def add_member():
    if "user" not in session:
        return jsonify({"success": False, "message": "Non autorisé"}), 401
    
    city = request.form["city"]
    if city == "Autre":
        city = request.form.get("custom_city", "")
    
    data = {
        "full_name": request.form["full_name"],
        "birth_date": request.form["birth_date"],
        "city": city,
        "join_date": request.form["join_date"],
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat()
    }
    
    if not is_valid_date(data["birth_date"]) or not is_valid_date(data["join_date"]):
        return jsonify({"success": False, "message": "Format de date invalide. Utilisez le format AAAA-MM-JJ."}), 400
    
    try:
        db.collection("members").add(data)
        return jsonify({"success": True, "message": "Membre ajouté avec succès"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": f"Erreur lors de l'ajout du membre: {str(e)}"}), 500

# Modifier un membre
@app.route("/update_member/<id>", methods=["POST"])
def update_member(id):
    if "user" not in session:
        return redirect(url_for("login"))
    
    data = {
        "full_name": request.form["full_name"],
        "birth_date": request.form["birth_date"],
        "city": request.form["city"],
        "join_date": request.form["join_date"],
        "updated_at": datetime.now().isoformat()
    }
    
    if not is_valid_date(data["birth_date"]) or not is_valid_date(data["join_date"]):
        flash("Format de date invalide. Utilisez le format AAAA-MM-JJ.")
        return redirect(url_for("home"))
    
    try:
        db.collection("members").document(id).set(data, merge=True)
        flash("Membre modifié avec succès")
    except Exception as e:
        flash(f"Erreur lors de la modification du membre: {str(e)}")
    
    return redirect(url_for("home"))

# Supprimer un membre
@app.route("/delete_member/<id>", methods=["DELETE"])
def delete_member(id):
    if "user" not in session:
        return jsonify({"success": False, "message": "Non autorisé"}), 401
    
    try:
        doc_ref = db.collection("members").document(id)
        if not doc_ref.get().exists:
            return jsonify({"success": False, "message": "Membre non trouvé"}), 404
            
        doc_ref.delete()
        return jsonify({"success": True, "message": "Membre supprimé avec succès"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": f"Erreur lors de la suppression: {str(e)}"}), 500

# Déconnexion
@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))

# Export Excel
@app.route("/export_excel")
def export_excel():
    if "user" not in session:
        return redirect(url_for("login"))

    # Récupérer les paramètres de filtre
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    city = request.args.get("city", "")
    
    # Récupération des données avec filtres si présents
    members_ref = db.collection("members")
    
    # Détecter si un filtre est appliqué
    is_filtered = year or month or city
    filter_description = "Tous les membres de SIS"
    
    if is_filtered:
        filter_parts = []
        if year:
            filter_parts.append(f"Année: {year}")
        if month:
            filter_parts.append(f"Mois: {month}")
        if city:
            filter_parts.append(f"Ville: {city}")
        filter_description = "Membres filtrés - " + ", ".join(filter_parts)
    
    # Préparation des données
    data = []
    for doc in members_ref.stream():
        member = doc.to_dict()
        # Appliquer les filtres côté serveur si nécessaire
        if is_filtered:
            member_year = member.get('join_date', '')[:4]
            member_month = member.get('join_date', '')[5:7]
            member_city = member.get('city', '').lower()
            
            if (year and year != member_year) or \
               (month and month != member_month) or \
               (city and city.lower() not in member_city):
                continue
                
        data.append({
            'Nom Complet': member.get('full_name', ''),
            'Date de Naissance': member.get('birth_date', ''),
            'Ville': member.get('city', ''),
            'Date Inscription': member.get('join_date', '')
        })

    # Création du fichier Excel avec style
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    import io

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Membres"

    # Ajouter le titre de l'export (fusion de cellules)
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = filter_description
    title_cell.font = Font(bold=True, size=14, color="4F81BD")
    title_cell.alignment = Alignment(horizontal="center")
    
    # Ajout des en-têtes (décalés d'une ligne)
    headers = ["Nom Complet", "Date de Naissance", "Ville", "Date Inscription"]
    ws.append(headers)

    # Remplissage des données
    for member in data:
        ws.append([member[h] for h in headers])

    # Style des en-têtes
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    for col in range(1, 5):
        cell = ws.cell(row=2, column=col)  # Notez le row=2 pour le décalage
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Style des cellules de données
    data_font = Font(name="Calibri", size=11)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=4):  # Commence à row=3
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left")

    # Ajustement des largeurs de colonnes
    column_widths = {
        'A': 30,  # Nom Complet
        'B': 20,  # Date de Naissance
        'C': 25,  # Ville
        'D': 20   # Date Inscription
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Ajouter un filtre automatique (décalé d'une ligne)
    ws.auto_filter.ref = f"A2:D{ws.max_row}"

    # Sauvegarde dans le buffer
    wb.save(output)
    output.seek(0)

    # Préparation de la réponse
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=membres_steps_into_space.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=10000, debug=True)
